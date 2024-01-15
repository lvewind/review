from bs4 import BeautifulSoup
import urllib.request
import time
import json
from openpyxl import load_workbook


class COVID_19:
    def __init__(self, template_file="template.xlsx", url = "https://voice.baidu.com/act/newpneumonia/newpneumonia/"):
        self.url = url
        self.template_file = template_file
        self.caseList = {}
        self.caseOutsideList = {}

    def get_case_list(self):
        user_agent = 'Mozilla/5.0 (Windows NT 6.1; Win64; x64)'
        headers = {'User-Agent': user_agent}
        req = urllib.request.Request(self.url, headers=headers)
        response = urllib.request.urlopen(req, timeout=10)
        html = response.read()
        soup = BeautifulSoup(html, "html.parser", from_encoding="utf-8")
        json_string = soup.find(id="captain-config").string
        data_dict = json.loads(json_string)
        self.caseList = data_dict.get("component")[0].get("caseList")
        # print(self.caseList)
        self.caseOutsideList = data_dict.get("component")[0].get("caseOutsideList")
        # print(self.caseOutsideList)

    def generate_excel_file(self):
        wb = load_workbook(filename=self.template_file)
        ws = wb.active
        # ws = wb_template.copy_worksheet(ws_template)
        # sheetnames = wb_template.sheetnames
        # ws_template = wb_template[sheetnames[0]]
        # print(ws_template)

        for i in range(5, 186):
            name = ws.cell(i, 2).value
            print(name)
            for item in self.caseList:
                if item.get("area") == name:
                    ws.cell(i, 3, int(item.get("confirmed")))
                    ws.cell(i, 4, int(item.get("curConfirm")))
                    ws.cell(i, 5, int(item.get("crued")))
                    ws.cell(i, 6, int(item.get("died")))
                    break
            else:
                for item in self.caseOutsideList:
                    if item.get("area") == name:
                        ws.cell(i, 3, int(item.get("confirmed")))
                        ws.cell(i, 4, int(item.get("curConfirm")))
                        ws.cell(i, 5, int(item.get("crued")))
                        ws.cell(i, 6, int(item.get("died")))
                        break
        localtime = time.localtime(time.time())
        new_file_name = str(localtime[0]) + str(localtime[1]) + str(localtime[2]) + str(localtime[3]) + str(localtime[4])
        ws.cell(3, 7, str(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
        ws["C186"] = "=SUM(C5:C185)"
        ws["D186"] = "=SUM(D5:D185)"
        ws["E186"] = "=SUM(E5:E185)"
        ws["F186"] = "=SUM(F5:F185)"

        wb.save(new_file_name + ".xlsx")
    
    def test_fun(self):
        pass

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    c19 = COVID_19()
    c19.get_case_list()
    c19.generate_excel_file()


